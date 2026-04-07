import pytest
import openpyxl

'''Skip, skipif and parameterized'''
# @pytest.mark.skip
# def test_01():
#     assert 'm' == 'm', "not matched"
#
#
# def test_02():
#     assert 'a' == 'a', "not matched"
#
#
# @pytest.mark.skipif(1!=1, reason='wrong')
# def test_03():
#     assert 1 != 0, 'matched'
#
#
# @pytest.mark.parametrize("a, b, expected", [
#     (1,2,3),
#     (2,3,5)
# ])
# def test_04(a, b, expected):
#     assert a+b == expected

'''Pytest Fixture'''
# @pytest.fixture(autouse=True)
# def greet():
#     print("Hello")
#     yield
#     print("Bye")
#
#
# def test_morning():
#     print("Good Morning")
#
#
# def test_evening():
#     print("Good Evening")

'''Data driven test'''
# import openpyxl

wb = openpyxl.Workbook()
sheetName = "sheet1"

if sheetName in wb.sheetnames:
    ws = wb[sheetName]
else:
    ws = wb.create_sheet(sheetName)

ws['A1'] = 'username'
ws['B1'] = 'password'
ws['C1'] = 'expected'

# wb.save('sample.xlsx')

ws.append(['user1', 'pass1', 'Epic sadface'])
ws.append(["standard_user", "secret_sauce", 'Products'])
ws.append(['locked_out_user', 'secret_sauce', 'Epic sadface'])
ws.append(['standard_user', 'secret_s', 'Epic sadface'])

wb.save('sample.xlsx')

